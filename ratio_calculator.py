"""Compute financial ratios for Palantir_Financials.xlsx.

This script reads the INCOME_STATEMENT, BALANCE_SHEET, STOCKHOLDERS_EQUITY,
and CASH_FLOW sheets, then writes a RATIOS worksheet with the requested
financial ratios.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook


@dataclass
class SheetConfig:
    name: str
    header_row: int
    label_col: int
    data_start_row: int
    data_end_row: int
    year_cols: List[int]


SHEET_CONFIGS = {
    "INCOME_STATEMENT": SheetConfig(
        name="INCOME_STATEMENT",
        header_row=15,
        label_col=2,
        data_start_row=16,
        data_end_row=42,
        year_cols=[3, 4, 5],
    ),
    "BALANCE_SHEET": SheetConfig(
        name="BALANCE_SHEET",
        header_row=14,
        label_col=2,
        data_start_row=17,
        data_end_row=55,
        year_cols=[3, 4],
    ),
    "STOCKHOLDERS_EQUITY": SheetConfig(
        name="STOCKHOLDERS_EQUITY",
        header_row=17,
        label_col=2,
        data_start_row=18,
        data_end_row=62,
        year_cols=[3, 4, 5, 6, 7, 8, 9, 10],
    ),
    "CASH_FLOW": SheetConfig(
        name="CASH_FLOW",
        header_row=15,
        label_col=2,
        data_start_row=17,
        data_end_row=68,
        year_cols=[3, 4, 5],
    ),
}


def _normalize_label(label: str) -> str:
    return " ".join(str(label).strip().lower().split())


def _read_table(ws, config: SheetConfig) -> Dict[str, Dict[int, float]]:
    years = []
    for col in config.year_cols:
        year_val = ws.cell(row=config.header_row, column=col).value
        if isinstance(year_val, (int, float)):
            years.append(int(year_val))
        else:
            years.append(None)

    table: Dict[str, Dict[int, float]] = {}
    for row in range(config.data_start_row, config.data_end_row + 1):
        label = ws.cell(row=row, column=config.label_col).value
        if label in (None, ""):
            continue
        label_key = _normalize_label(label)
        row_values: Dict[int, float] = {}
        for col, year in zip(config.year_cols, years):
            if year is None:
                continue
            value = ws.cell(row=row, column=col).value
            if value in (None, ""):
                continue
            try:
                row_values[year] = float(value)
            except (TypeError, ValueError):
                continue
        if row_values:
            table[label_key] = row_values
    return table


def _find_value(
    table: Dict[str, Dict[int, float]],
    label_variants: Iterable[str],
    year: int,
) -> Optional[float]:
    for variant in label_variants:
        key = _normalize_label(variant)
        if key in table and year in table[key]:
            return table[key][year]
    # fallback: substring search
    for key, values in table.items():
        for variant in label_variants:
            if _normalize_label(variant) in key and year in values:
                return values[year]
    return None


def _average(current: Optional[float], prior: Optional[float]) -> Optional[float]:
    if current is None or prior is None:
        return None
    return (current + prior) / 2


def _safe_divide(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator


def build_ratios(workbook_path: str, output_path: Optional[str] = None) -> None:
    wb = load_workbook(workbook_path, data_only=True)
    tables = {}
    for name, config in SHEET_CONFIGS.items():
        ws = wb[config.name]
        tables[name] = _read_table(ws, config)

    income = tables["INCOME_STATEMENT"]
    balance = tables["BALANCE_SHEET"]

    # Use years from income statement where balance sheet has matching year data.
    years = sorted(
        set(next(iter(income.values())).keys()) & set(next(iter(balance.values())).keys())
    )

    ratios: Dict[str, Dict[int, Optional[float]]] = {
        "A/R Turnover": {},
        "Inventory Turnover": {},
        "A/P Turnover": {},
        "PPE Turnover": {},
        "Asset Turnover": {},
        "Return on Assets": {},
        "Cash Conversion Cycle": {},
        "Return on Equity": {},
        "Return on Common Equity": {},
        "Net Profit Margin": {},
        "Leverage": {},
    }

    for year in years:
        prior_year = year - 1
        revenue = _find_value(income, ["Revenue"], year)
        cost_of_revenue = _find_value(income, ["Cost of revenue"], year)
        net_income = _find_value(income, ["Net income (loss)"], year)
        net_income_common = _find_value(
            income,
            ["Net income (loss) attributable to common stockholders"],
            year,
        )

        accounts_receivable = _find_value(balance, ["Accounts receivable, net"], year)
        accounts_receivable_prior = _find_value(balance, ["Accounts receivable, net"], prior_year)
        inventory = _find_value(balance, ["Inventory"], year)
        inventory_prior = _find_value(balance, ["Inventory"], prior_year)
        accounts_payable = _find_value(balance, ["Accounts payable"], year)
        accounts_payable_prior = _find_value(balance, ["Accounts payable"], prior_year)
        ppe = _find_value(balance, ["Property and equipment, net"], year)
        ppe_prior = _find_value(balance, ["Property and equipment, net"], prior_year)
        total_assets = _find_value(balance, ["Total assets"], year)
        total_assets_prior = _find_value(balance, ["Total assets"], prior_year)
        total_equity = _find_value(balance, ["Total equity"], year)
        total_equity_prior = _find_value(balance, ["Total equity"], prior_year)
        total_stockholders_equity = _find_value(
            balance,
            ["Total Palantir's stockholders' equity"],
            year,
        )
        total_stockholders_equity_prior = _find_value(
            balance,
            ["Total Palantir's stockholders' equity"],
            prior_year,
        )

        avg_ar = _average(accounts_receivable, accounts_receivable_prior)
        avg_inventory = _average(inventory, inventory_prior)
        avg_ap = _average(accounts_payable, accounts_payable_prior)
        avg_ppe = _average(ppe, ppe_prior)
        avg_assets = _average(total_assets, total_assets_prior)
        avg_equity = _average(total_equity, total_equity_prior)
        avg_common_equity = _average(total_stockholders_equity, total_stockholders_equity_prior)

        ar_turnover = _safe_divide(revenue, avg_ar)
        inventory_turnover = _safe_divide(cost_of_revenue, avg_inventory)
        ap_turnover = _safe_divide(cost_of_revenue, avg_ap)
        ppe_turnover = _safe_divide(revenue, avg_ppe)
        asset_turnover = _safe_divide(revenue, avg_assets)
        roa = _safe_divide(net_income, avg_assets)

        dso = _safe_divide(365, ar_turnover)
        dio = _safe_divide(365, inventory_turnover)
        dpo = _safe_divide(365, ap_turnover)
        ccc = None
        if dso is not None and dpo is not None:
            ccc = dso - dpo if dio is None else dso + dio - dpo

        roe = _safe_divide(net_income, avg_equity)
        roce = _safe_divide(net_income_common, avg_common_equity)
        npm = _safe_divide(net_income, revenue)
        leverage = _safe_divide(avg_assets, avg_equity)

        ratios["A/R Turnover"][year] = ar_turnover
        ratios["Inventory Turnover"][year] = inventory_turnover
        ratios["A/P Turnover"][year] = ap_turnover
        ratios["PPE Turnover"][year] = ppe_turnover
        ratios["Asset Turnover"][year] = asset_turnover
        ratios["Return on Assets"][year] = roa
        ratios["Cash Conversion Cycle"][year] = ccc
        ratios["Return on Equity"][year] = roe
        ratios["Return on Common Equity"][year] = roce
        ratios["Net Profit Margin"][year] = npm
        ratios["Leverage"][year] = leverage

    # Create or replace RATIOS worksheet
    if "RATIOS" in wb.sheetnames:
        del wb["RATIOS"]
    ws_ratios = wb.create_sheet("RATIOS")

    ws_ratios.cell(row=1, column=1, value="Ratio")
    for col_idx, year in enumerate(years, start=2):
        ws_ratios.cell(row=1, column=col_idx, value=year)

    for row_idx, ratio_name in enumerate(ratios.keys(), start=2):
        ws_ratios.cell(row=row_idx, column=1, value=ratio_name)
        for col_idx, year in enumerate(years, start=2):
            ws_ratios.cell(row=row_idx, column=col_idx, value=ratios[ratio_name].get(year))

    wb.save(output_path or workbook_path)


if __name__ == "__main__":
    build_ratios("Palantir_Financials.xlsx")
